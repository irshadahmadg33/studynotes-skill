#!/usr/bin/env python3
"""
Save quiz results to an Excel file.

This script creates an Excel file with quiz results including:
- Question number
- User's answer
- Correct answer
- Right/Wrong status
- Overall score

Usage:
    python save_quiz_results.py <output_file> <quiz_data_json>
"""

import sys
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl library is required. Install it with: pip install openpyxl")
    sys.exit(1)


def create_quiz_results_excel(output_file, quiz_data):
    """
    Create an Excel file with quiz results.

    Args:
        output_file (str): Path to the output Excel file
        quiz_data (dict): Dictionary containing quiz results with keys:
            - topic (str): The quiz topic
            - questions (list): List of question dictionaries with:
                - question_num (int)
                - user_answer (str)
                - correct_answer (str)
                - is_correct (bool)
            - score (dict): Dictionary with 'correct' and 'total' keys
            - timestamp (str, optional): Quiz completion timestamp
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Results"

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Quiz Results - {quiz_data.get('topic', 'General Quiz')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    # Add timestamp if available
    timestamp = quiz_data.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Completed: {timestamp}"
    ws['A2'].alignment = center_align

    # Add score summary
    score = quiz_data.get('score', {})
    correct = score.get('correct', 0)
    total = score.get('total', 0)
    percentage = (correct / total * 100) if total > 0 else 0

    ws.merge_cells('A3:E3')
    ws['A3'] = f"Score: {correct}/{total} ({percentage:.1f}%)"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = center_align

    # Add headers
    headers = ['Question #', 'Your Answer', 'Correct Answer', 'Status', 'Result']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Add question results
    questions = quiz_data.get('questions', [])
    for idx, q in enumerate(questions, start=6):
        ws.cell(row=idx, column=1, value=q['question_num'])
        ws.cell(row=idx, column=2, value=q['user_answer'])
        ws.cell(row=idx, column=3, value=q['correct_answer'])
        ws.cell(row=idx, column=4, value='✓' if q['is_correct'] else '✗')
        ws.cell(row=idx, column=5, value='Correct' if q['is_correct'] else 'Incorrect')

        # Apply styling
        for col in range(1, 6):
            cell = ws.cell(row=idx, column=col)
            cell.alignment = center_align
            cell.border = border
            if q['is_correct']:
                cell.fill = correct_fill
            else:
                cell.fill = incorrect_fill

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Save the workbook
    wb.save(output_file)
    print(f"Quiz results saved to: {output_file}")


def main():
    if len(sys.argv) < 3:
        print("Usage: python save_quiz_results.py <output_file> <quiz_data_json>")
        sys.exit(1)

    output_file = sys.argv[1]
    quiz_data_json = sys.argv[2]

    try:
        quiz_data = json.loads(quiz_data_json)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON data: {e}")
        sys.exit(1)

    try:
        create_quiz_results_excel(output_file, quiz_data)
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
